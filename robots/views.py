import json

from datetime import datetime as dt, timedelta

from django.core import serializers
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Count
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt

from openpyxl import Workbook

from .models import Robot


@csrf_exempt
def ApiView(request):
    """API-view for requests."""

    if request.method == 'GET':
        try:
            data = json.loads(serializers.serialize('json',
                                                    Robot.objects.all()))
            new_data = []
            for dct in data:
                new_data.append(dct.get('fields'))
            return JsonResponse(new_data, encoder=DjangoJSONEncoder,
                                safe=False)
        except Exception as e:
            raise ValueError(f'{e}: the form is not valid.')

    if request.method == 'POST':
        body = json.loads(request.body.decode('utf-8'))
        new_item = Robot.objects.create(serial=body['serial'],
                                        model=body['model'],
                                        version=body['version'],
                                        created=body['created']
                                        )
        try:
            data = json.loads(serializers.serialize('json', new_item))
            return JsonResponse(data, encoder=DjangoJSONEncoder, safe=False)
        except Exception as e:
            raise ValueError(f'{e}: the form is not valid.')


@csrf_exempt
def ApiIdView(request, id):
    """API-view for particular requests."""

    if request.method == 'GET':
        data = json.loads(serializers.serialize('json',
                                                Robot.objects.filter(id=id)))
        new_data = []
        for dct in data:
            new_data.append(dct.get('fields'))
        return JsonResponse(new_data, encoder=DjangoJSONEncoder,
                            safe=False)

    if request.method == 'PUT':
        body = json.loads(request.body.decode('utf-8'))
        Robot.objects.filter(
                             pk=id
                            ).update(
                                     serial=body['serial'],
                                     model=body['model'],
                                     version=body['version'],
                                     created=body['created']
                                     )
        updated_item = Robot.objects.filter(id=id)
        try:
            data = json.loads(serializers.serialize('json', updated_item))
            return JsonResponse(data, encoder=DjangoJSONEncoder, safe=False)
        except Exception as e:
            raise ValueError(f'{e}: the form is not valid.')

    if request.method == 'DELETE':
        Robot.objects.get(id=id).delete()
        data = json.loads(serializers.serialize('json', Robot.objects.all()))
        return JsonResponse(data, encoder=DjangoJSONEncoder, safe=False)


class SummaryReportView(View):

    def get(self, request):
        current_date = dt.now()
        start_date = current_date - timedelta(days=7)

        robots = Robot.objects.filter(created__gte=start_date
                                      ).values('model', 'version'
                                               ).annotate(count=Count('id'))

        wb = Workbook()

        for model_data in robots:
            model = model_data['model']
            version = model_data['version']
            count = model_data['count']

            sheet = (wb[model] if model in wb.sheetnames
                     else wb.create_sheet(model))

            sheet['A1'] = 'Модель'
            sheet["B1"] = 'Версия'
            sheet["C1"] = 'Количество за неделю'

            last_row = sheet.max_row + 1

            sheet.cell(row=last_row, column=1, value=model)
            sheet.cell(row=last_row, column=2, value=version)
            sheet.cell(row=last_row, column=3, value=count)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition':
                     'attachment; filename=summary_report.xlsx'
                     }
            )
        wb.save(response)

        return response
